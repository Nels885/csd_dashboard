import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from utils.microsoft_format import CsvFormat, ExcelFormat, pd
from utils.file.export import ExportExcel


class CsvSuptech(CsvFormat):
    COLS = {"DATE": "date", "QUI": "user", "XELON": "xelon", "ITEM": "item", "TIME": "time", "INFO": "info",
            "RMQ": "rmq"}
    COLS_DATE = {'date': "%d/%m/%Y"}

    def __init__(self, file, columns=None):
        """
        Initialize ExcelSqualaetp class
        :param file:
            excel file to process
        :param sheet_index:
            Sheet index to be processed from excel file
        :param columns:
            Number of the last column to be processed
        """
        try:
            super(CsvSuptech, self).__init__(file, columns, dtype=str, usecols=self.COLS.keys())
            self._columns_rename(self.COLS)
            self.sheet.replace({"": None}, inplace=True)
            self._date_converter(self.COLS_DATE)
            self.error = False
        except pd.errors.EmptyDataError:
            self.error = True

    def read(self):
        """
        Formatting data from the Raspeedi excel file
        :return:
            list of dictionnaries that represents the data in the sheet
        """
        data = []
        if not self.error:
            for line in range(self.nrows):
                data.append(dict(self.sheet.loc[line].dropna()))
        return data


class ExcelSuptech(ExcelFormat):
    """## Read data in Excel file for LOG_SUPTECH.xlsx ##"""
    COLS = {"DATE": "date", "QUI": "user", "XELON": "xelon", "ITEM": "item", "TIME": "time", "INFO": "info",
            "RMQ": "rmq", "ACTION/RETOUR": "action"}
    COLS_DATE = {'date': "%d/%m/%Y %H:%M:%S"}

    def __init__(self, file, sheet_name=0, columns=None):
        """
        Initialize ExcelSqualaetp class
        :param file:
            excel file to process
        :param sheet_name:
            Sheet index to be processed from excel file
        :param columns:
            Number of the last column to be processed
        """
        super(ExcelSuptech, self).__init__(file, sheet_name, columns)
        self._columns_rename(self.COLS)
        self.sheet.replace({"": None}, inplace=True)
        self._date_converter(self.COLS_DATE)

    def read(self):
        """
        Extracting data for the Suptech table form the Database
        :return:
            list of dictionnaries that represents the data for Suptech table
        """
        data = []
        for line in range(self.nrows):
            row = self.sheet.loc[line]  # get the data in the ith row
            data.append(dict(row.dropna()))
        return data


class ExportExcelSuptech(ExportExcel):

    def __init__(self, values_list, filename, header, novalue="#"):
        super().__init__(values_list, filename, header, novalue)

    def _xlsx_writer(self, response):
        """ Formatting data in Excel 2010 format """
        wb = openpyxl.Workbook()

        # Get active worksheet/tab
        ws = wb.active
        ws.title = 'Feuille 1'
        col_dimensions = {'A': 20, 'B': 12, 'C': 13, 'D': 25, 'E': 8, 'F': 80, 'G': 60, 'H': 60}
        for key, value in col_dimensions.items():
            ws.column_dimensions[key].width = value
        ws['F2'].alignment = Alignment(wrap_text=True)

        # Sheet header, first row
        row_num = 1

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(self.header, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="b2b2b2", end_color="b2b2b2", fill_type="solid")
            cell.border = thin_border
            cell.value = column_title

        # Iterate though all values
        for query in self.valueSet:
            row_num += 1
            query = tuple([self._html_to_string(_) if isinstance(_, str) else _ for _ in query])
            query = self._query_format(query)

            # Assign the data  for each cell of the row
            for col_num, cell_value in enumerate(query, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.alignment = Alignment(wrap_text=True, vertical='center')
                cell.border = thin_border
            ws.row_dimensions[row_num].height = 75

        wb.save(response)
        return response
