import re
import csv
import xlwt
import openpyxl
from openpyxl.styles import Font
import datetime
from . import os
import shutil
import logging
from html.parser import HTMLParser

from django.shortcuts import HttpResponse
# from django.utils.safestring import SafeString

from psa.models import Corvet
from utils.conf import XML_CORVET_PATH, TAG_XELON_PATH, TAG_XELON_LOG_PATH

# Get an instance of a logger
logger = logging.getLogger('command')


class HTMLFilter(HTMLParser):
    text = ""

    def handle_data(self, data):
        self.text += data


def xml_corvet_file(instance, data, vin):
    try:
        xelon_nb = instance.numero_de_dossier
        os.makedirs(XML_CORVET_PATH, exist_ok=True)
        file = os.path.join(XML_CORVET_PATH, xelon_nb + ".xml")
        if not os.path.isfile(file):
            with open(file, "w", encoding='utf-8') as f:
                f.write(str(data))
        else:
            logger.warning("{} File exists.".format(xelon_nb))
    except Corvet.DoesNotExist:
        logger.warning("Xelon number not found")


class Calibre:
    """
    Class allowing the processing of calibration files for Xelon unlocking
    """

    def __init__(self, *args):
        self.paths = list(args)

    def file(self, xelon, comments, user):
        """
        Generate xelon unlock file
        :param xelon: Xelon number
        :param comments: User comment
        """
        if xelon != 'A123456789':
            for path in self.paths:
                file = os.path.join(path, xelon + ".txt")
                os.makedirs(path, exist_ok=True)
                if not os.path.isfile(file):
                    with open(file, "w", encoding='utf-8') as f:
                        f.write("Configuration produit effectuée par {}\r\n{}".format(user, comments))
                else:
                    logger.warning("%s File exists.", xelon)
                    return False
        return True

    def check(self, xelon):
        """
        Check if the file exists
        :param xelon: Xelon number
        """
        file = os.path.join(self.paths[0], xelon + ".txt")
        if os.path.isfile(file):
            return True
        return False


calibre = Calibre(TAG_XELON_PATH, TAG_XELON_LOG_PATH)


class ExportExcel:
    """ class for exporting data in CSV format """

    def __init__(self, values_list, filename, header, novalue="#"):
        self.date = datetime.datetime.now()
        self.filename, self.excelType = self._file_format(filename)
        self.header = header
        self.noValue = novalue
        self.valueSet = values_list

    def http_response(self):
        """ Creation http response """
        if self.excelType == "csv":
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="{}_{}.{}"'.format(
                self.filename, self.date.strftime("%y-%m-%d_%H-%M"), self.excelType
            )
            self._csv_writer(response)
        elif self.excelType == "xlsx":
            response = HttpResponse(content_type='application/ms_excel')
            response['Content-Disposition'] = 'attachment; filename="{}_{}.{}"'.format(
                self.filename, self.date.strftime("%y-%m-%d_%H-%M"), self.excelType
            )
            self._xlsx_writer(response)
        else:
            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachement; filename="{}_{}.{}'.format(
                self.filename, self.date.strftime("%y-%m-%d_%H-%M"), self.excelType
            )
            self._xlsx_writer(response)
        return response

    def file(self, path, copy=True):
        """
        Creation file
        :return True if the file is read-only
        """
        file = os.path.join(path, "{}.{}".format(self.filename, self.excelType))
        if copy:
            self._file_yesterday(path, file)
        try:
            if self.excelType == "csv":
                with open(file, 'w+', newline='', encoding='utf-8-sig') as f:
                    if self.excelType == "csv":
                        self._csv_writer(f)
            elif self.excelType == "xlsx":
                self._xlsx_writer(file)
            else:
                self._xlsx_writer(file)
            return False
        except OSError:
            logger.warning('{} File is read-only.'.format(file))
        return True

    def _csv_writer(self, response):
        """ Formatting data in CSV format """
        writer = csv.writer(response, delimiter=';', lineterminator=';\r\n')
        writer.writerow(self.header)

        for i, query in enumerate(self.valueSet):
            query = tuple([self._html_to_string(_, r'[;,]') if isinstance(_, str) else _ for _ in query])
            query = self._query_format(query)
            writer.writerow(query)

    def _xls_writer(self, response):
        """ Formatting data in Excel format """
        xldoc = xlwt.Workbook(encoding='utf-8')
        sheet = xldoc.add_sheet('Feuille 1')

        # Sheet header, first row
        row_num = 0

        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        for col_num in range(len(self.header)):
            sheet.write(row_num, col_num, self.header[col_num], font_style)

        # Sheet body, remaining rows
        font_style = xlwt.XFStyle()

        for i, query in enumerate(self.valueSet):
            query = tuple([self._html_to_string(_) if isinstance(_, str) else _ for _ in query])
            query = self._query_format(query)
            row_num += 1
            for col_num in range(len(query)):
                sheet.write(row_num, col_num, query[col_num], font_style)

        xldoc.save(response)
        return response

    def _xlsx_writer(self, response):
        """ Formatting data in Excel 2010 format """
        wb = openpyxl.Workbook()

        # Get active worksheet/tab
        ws = wb.active
        ws.title = 'Feuille 1'

        # Sheet header, first row
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(self.header, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = Font(bold=True)
            cell.value = column_title

        # Iterate though all values
        for query in self.valueSet:
            row_num += 1
            query = tuple([self._html_to_string(_) if isinstance(_, str) else _ for _ in query])
            query = self._query_format(query)

            # Assign the data  for each cell of the row
            for col_num, cell_value in enumerate(query, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = cell_value

        wb.save(response)
        return response

    def _file_yesterday(self, path, file):
        """ Creation of the backup file d-1 """
        yesterday = datetime.date.today() - datetime.timedelta(days=1)
        if os.path.isfile(file):
            file_date = datetime.datetime.fromtimestamp(os.path.getmtime(file)).date()
            if file_date >= yesterday:
                shutil.copyfile(file, os.path.join(path, "{}J-1.{}".format(self.filename, self.excelType)))

    def _query_format(self, query):
        query = tuple([_.strftime("%d/%m/%Y %H:%M:%S") if isinstance(_, datetime.date) else _ for _ in query])
        query = tuple([self.noValue if not value else value for value in query])
        return query

    @staticmethod
    def _html_to_string(value, re_sub=None):
        f = HTMLFilter()
        f.feed(value)
        if re_sub:
            return re.sub(re_sub, ' ', f.text)
        else:
            return f.text

    @staticmethod
    def _file_format(filename):
        file_list = filename.split('.')
        filename, extension = file_list[0], 'csv'
        if len(file_list) > 1:
            extension = file_list[-1]
        return filename, extension
