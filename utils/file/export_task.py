import tempfile
import datetime
import csv
import xlwt
import os.path
import shutil

from django.utils import timezone
from celery_progress.backend import ProgressRecorder
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from sbadmin.tasks.base import BaseTask
from utils.file.export import HTMLFilter, re


class ExportExcelTask(BaseTask):
    name = "ExportExcelTask"

    def __init__(self, *args, **kwargs) -> None:
        super().__init__(*args, **kwargs)
        self.sheetName = kwargs.get('sheet_name', 'Feuil1')
        self.noValue = kwargs.get('novalue', "#")
        self.textCols = kwargs.get('text_cols', [])
        self.header = kwargs.get('header', [])

    @staticmethod
    def copy_and_get_copied_path():
        destination_path = "%s" % (tempfile.gettempdir())
        return destination_path

    @staticmethod
    def get_header_fields(prod_list):
        header = [value_tuple[0] for value_tuple in prod_list]
        fields = [value_tuple[1] for value_tuple in prod_list]
        return header, fields

    def file(self, filename, excel_type, values_list):
        """
        Creation file
        :param filename: file name
        :param excel_type: type of Excel file (xls, xlsx, csv)
        :param values_list: List of values to include in the file
        :return Destination path to file
        """
        path = self.copy_and_get_copied_path()
        destination_path = os.path.join(
            path, f"{filename.replace('/', '-')}_{timezone.now().strftime('%y-%m-%d_%H-%M')}.{excel_type}")
        if excel_type == "csv":
            self._create_csv(destination_path, values_list)
        elif excel_type == "xls":
            xldoc = xlwt.Workbook(encoding='utf-8-sig')
            xldoc = self._create_xls(xldoc, values_list)
            xldoc.save(destination_path)
        else:
            workbook = Workbook()
            workbook = self._create_xlsx(workbook, values_list)
            workbook.save(filename=destination_path)
        return destination_path

    def _create_xlsx(self, workbook: Workbook, values_list):
        """ Formatting data in Excel 2010 format """
        progress_recorder = ProgressRecorder(self)
        # Get active worksheet/tab
        workbook.encoding = 'utf-8-sig'
        ws = workbook.active
        ws.title = self.sheetName

        # Sheet header, first row
        row_num = 1
        total_record = len(values_list)

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(self.header, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = Font(bold=True)
            cell.value = column_title

        # Iterate though all values
        for query in values_list:
            row_num += 1
            query = self._query_format(query)

            # Assign the data  for each cell of the row
            for col_num, cell_value in enumerate(query, 1):
                cell = ws.cell(row=row_num, column=col_num)
                if col_num in self.textCols:
                    cell.alignment = Alignment(wrapText=True)
                cell.value = cell_value
            progress_recorder.set_progress(row_num + 1, total=total_record, description="Inserting record into row")
        return workbook

    def _create_xls(self, xldoc: xlwt.Workbook, values_list):
        """ Formatting data in Excel format """
        progress_recorder = ProgressRecorder(self)
        sheet = xldoc.add_sheet(self.sheetName)

        # Sheet header, first row
        row_num = 0
        total_record = len(values_list)

        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        for col_num in range(len(self.header)):
            sheet.write(row_num, col_num, self.header[col_num], font_style)

        # Sheet body, remaining rows
        font_style = xlwt.XFStyle()

        for query in values_list:
            row_num += 1
            query = self._query_format(query)

            for col_num in range(len(query)):
                sheet.write(row_num, col_num, query[col_num], font_style)
            progress_recorder.set_progress(row_num + 1, total=total_record, description="Inserting record into row")
        return xldoc

    def _create_csv(self, filename, values_list):
        """ Formatting data in CSV format """
        progress_recorder = ProgressRecorder(self)
        total_record = len(values_list)
        with open(filename, 'w+', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f, delimiter=';', lineterminator=';\r\n')
            writer.writerow(self.header)

            for row_num, query in enumerate(values_list):
                query = tuple([self._html_to_string(_, r'[;,]') if isinstance(_, str) else _ for _ in query])
                query = self._query_format(query)
                writer.writerow(query)
                progress_recorder.set_progress(row_num + 1, total=total_record, description="Inserting record into row")

    def _query_format(self, query):
        query = tuple([self._html_to_string(_) if isinstance(_, str) else _ for _ in query])
        query = tuple([self._timestamp_to_string(_) for _ in query])
        query = tuple([self.noValue if not value else value for value in query])
        return query

    def _file_yesterday(self, path, file):
        """ Creation of the backup file d-1 """
        yesterday = timezone.now() - timezone.timedelta(days=1)
        if os.path.isfile(file):
            file_date = timezone.datetime.fromtimestamp(os.path.getmtime(file)).date()
            if file_date >= yesterday:
                shutil.copyfile(file, os.path.join(path, "{}J-1.{}".format(self.filename, self.excelType)))

    @staticmethod
    def _timestamp_to_string(value):
        if isinstance(value, datetime.datetime):
            value = value.strftime("%d/%m/%Y %H:%M:%S").replace(" 00:00:00", "")
        elif isinstance(value, datetime.date):
            value = value.strftime("%d/%m/%Y")
        elif isinstance(value, datetime.time):
            value = value.strftime("%H:%M:%S")
        return value

    @staticmethod
    def _html_to_string(value, re_sub=None):
        f = HTMLFilter()
        f.feed(value)
        if re_sub:
            return re.sub(re_sub, ' ', f.text)
        else:
            return f.text


class ExportExcel(ExportExcelTask):
    name = "ExportExcel"

    def __init__(self, *args, **kwargs) -> None:
        super().__init__(*args, **kwargs)

    def _create_xlsx(self, workbook: Workbook, values_list):
        """ Formatting data in Excel 2010 format """
        # Get active worksheet/tab
        workbook.encoding = 'utf-8-sig'
        ws = workbook.active
        ws.title = self.sheetName

        # Sheet header, first row
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(self.header, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = Font(bold=True)
            cell.value = column_title

        # Iterate though all values
        for query in values_list:
            row_num += 1
            query = self._query_format(query)

            # Assign the data  for each cell of the row
            for col_num, cell_value in enumerate(query, 1):
                cell = ws.cell(row=row_num, column=col_num)
                if col_num in self.textCols:
                    cell.alignment = Alignment(wrapText=True)
                cell.value = cell_value
        return workbook

    def _create_xls(self, xldoc: xlwt.Workbook, values_list):
        """ Formatting data in Excel format """
        sheet = xldoc.add_sheet(self.sheetName)

        # Sheet header, first row
        row_num = 0

        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        for col_num in range(len(self.header)):
            sheet.write(row_num, col_num, self.header[col_num], font_style)

        # Sheet body, remaining rows
        font_style = xlwt.XFStyle()

        for query in values_list:
            row_num += 1
            query = self._query_format(query)

            for col_num in range(len(query)):
                sheet.write(row_num, col_num, query[col_num], font_style)
        return xldoc

    def _create_csv(self, filename, values_list):
        """ Formatting data in CSV format """
        with open(filename, 'w+', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f, delimiter=';', lineterminator=';\r\n')
            writer.writerow(self.header)

            for row_num, query in enumerate(values_list):
                query = tuple([self._html_to_string(_, r'[;,]') if isinstance(_, str) else _ for _ in query])
                query = self._query_format(query)
                writer.writerow(query)
