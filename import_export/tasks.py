import re
import os.path
import datetime

from sbadmin import celery_app
from openpyxl import Workbook

from .utils import extract_corvet, extract_ecu, extract_reman, extract_tools
from utils.file.export_task import ExportExcelTask
from psa.models import Multimedia
from psa.templatetags.corvet_tags import get_corvet


""" source: https://github.com/ebysofyan/django-celery-progress-sample """


class ExportCorvetIntoExcelTask(ExportExcelTask):
    COL_CORVET = {
        'corvet__donnee_ligne_de_produit': 'DON_LIN_PROD', 'corvet__donnee_silhouette': 'DON_SIL',
        'corvet__donnee_genre_de_produit': 'DON_GEN_PROD', 'corvet__attribut_dhb': 'ATT_DHB',
        'corvet__attribut_dlx': 'ATT_DLX', 'corvet__attribut_dun': 'ATT_DUN', 'corvet__attribut_dym': 'ATT_DYM',
        'corvet__attribut_dyr': 'ATT_DYR', 'corvet__donnee_moteur': 'DON_MOT'
    }

    def __init__(self, *args, **kwargs) -> None:
        super().__init__(*args, **kwargs)

    def _query_format(self, query):
        data_list = [value for value in query]
        data_list = self.get_multimedia_display(data_list)
        query = self.get_corvet_display(data_list)
        query = tuple([_.strftime("%d/%m/%Y %H:%M:%S") if isinstance(_, datetime.date) else _ for _ in query])
        query = tuple([self.noValue if not value else value for value in query])
        return query

    def get_multimedia_display(self, data_list):
        if 'corvet__prods__btel__name' in self.fields:
            position = self.fields.index('corvet__prods__btel__name')
            for prod in Multimedia.PRODUCT_CHOICES:
                if prod[0] == data_list[position]:
                    data_list[position] = prod[1]
                    break
        return data_list

    def get_corvet_display(self, data_list):
        for field, arg in self.COL_CORVET.items():
            if field in self.fields:
                position = self.fields.index(field)
                if data_list[position]:
                    if 'vin' in self.fields and arg == 'DON_LIN_PROD':
                        if re.match(r'^[V][FR]7\w{14}$', str(data_list[self.fields.index('vin')])):
                            arg = 'DON_LIN_PROD 1'
                        else:
                            arg = 'DON_LIN_PROD 0'
                    data_list[position] = f"{data_list[position]} - {get_corvet(data_list[position], arg)}"
        return data_list

    def run(self, *args, **kwargs):
        path = self.copy_and_get_copied_path()
        excel_type = kwargs.pop('excel_type', 'xlsx')
        product = kwargs.pop('product', 'bsi')
        vin_list = kwargs.pop('vin_list', None)
        if vin_list is None:
            filename = f"{product}_{self.date.strftime('%y-%m-%d_%H-%M')}"
            self.header, self.fields, values_list = extract_corvet(product)
        else:
            filename = f"ecu_{self.date.strftime('%y-%m-%d_%H-%M')}"
            self.header, self.fields, values_list = extract_ecu(vin_list)
        destination_path = os.path.join(path, f"{filename}.{excel_type}")
        workbook = Workbook()
        workbook = self.create_workbook(workbook, self.header, values_list)
        workbook.save(filename=destination_path)
        return {
            "detail": "Successfully export CORVET",
            "data": {
                "outfile": destination_path
            }
        }


class ExportRemanIntoExcelTask(ExportExcelTask):

    def __init__(self, *args, **kwargs) -> None:
        super().__init__(*args, **kwargs)

    def run(self, *args, **kwargs):
        path = self.copy_and_get_copied_path()
        excel_type = kwargs.pop('excel_type', 'xlsx')
        model = kwargs.pop('table', 'bsi')
        filename = f"{model}_{self.date.strftime('%y-%m-%d_%H-%M')}"
        self.header, self.fields, values_list = extract_reman(model)
        destination_path = os.path.join(path, f"{filename}.{excel_type}")
        workbook = Workbook()
        workbook = self.create_workbook(workbook, self.header, values_list)
        workbook.save(filename=destination_path)
        return {
            "detail": "Successfully export REMAN",
            "data": {
                "outfile": destination_path
            }
        }


class ExportToolsIntoExcelTask(ExportExcelTask):

    def __init__(self, *args, **kwargs) -> None:
        super().__init__(*args, **kwargs)
        self.noValue = ""

    def run(self, *args, **kwargs):
        path = self.copy_and_get_copied_path()
        excel_type = kwargs.pop('excel_type', 'xlsx')
        model = kwargs.pop('table', 'suptech')
        filename = f"{model}_{self.date.strftime('%y-%m-%d_%H-%M')}"
        self.header, self.fields, values_list = extract_tools(model)
        destination_path = os.path.join(path, f"{filename}.{excel_type}")
        workbook = Workbook()
        workbook = self.create_workbook(workbook, self.header, values_list)
        workbook.save(filename=destination_path)
        return {
            "detail": "Successfully export TOOLS",
            "data": {
                "outfile": destination_path
            }
        }


@celery_app.task(bind=True, base=ExportCorvetIntoExcelTask)
def export_corvet_task(self, *args, **kwargs):
    return super(type(self), self).run(*args, **kwargs)


@celery_app.task(bind=True, base=ExportRemanIntoExcelTask)
def export_reman_task(self, *args, **kwargs):
    return super(type(self), self).run(*args, **kwargs)


@celery_app.task(bind=True, base=ExportToolsIntoExcelTask)
def export_tools_task(self, *args, **kwargs):
    return super(type(self), self).run(*args, **kwargs)
